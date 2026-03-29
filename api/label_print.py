"""
Vercel Python ランタイム用ラベル印刷（Node 側に python3 が無い環境向け）。
環境変数 REM_LABEL_PRINT_INTERNAL_SECRET とヘッダー x-rem-label-secret が一致する場合のみ実行。
"""
import io
import json
import math
import os
import re
from http.server import BaseHTTPRequestHandler
from pathlib import Path

from openpyxl import load_workbook

PREFS = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県", "茨城県", "栃木県", "群馬県",
    "埼玉県", "千葉県", "東京都", "神奈川県", "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県", "福岡県",
    "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]

SLOTS = [
    {"zip": "B2", "addr": "B3", "name": "C8", "sama": "D8"},
    {"zip": "G2", "addr": "G3", "name": "H8", "sama": "I8"},
    {"zip": "B12", "addr": "B13", "name": "C18", "sama": "D18"},
    {"zip": "G12", "addr": "G13", "name": "H18", "sama": "I18"},
    {"zip": "B22", "addr": "B23", "name": "C28", "sama": "D28"},
    {"zip": "G22", "addr": "G23", "name": "H28", "sama": "I28"},
    {"zip": "B32", "addr": "B33", "name": "C38", "sama": "D38"},
    {"zip": "G32", "addr": "G33", "name": "H38", "sama": "I38"},
    {"zip": "B42", "addr": "B43", "name": "C48", "sama": "D48"},
    {"zip": "G42", "addr": "G43", "name": "H48", "sama": "I48"},
    {"zip": "B52", "addr": "B53", "name": "C58", "sama": "D58"},
    {"zip": "G52", "addr": "G53", "name": "H58", "sama": "I58"},
]


def t(v):
    if v is None:
        return ""
    return str(v).strip()


def pref_label(raw):
    s = t(raw)
    if not s:
        return ""
    if s in PREFS:
        return s
    try:
        n = float(s)
        i = int(n)
        if 1 <= i <= len(PREFS):
            return PREFS[i - 1]
    except Exception:
        pass
    return s


def format_zip(raw):
    digits = re.sub(r"\D", "", t(raw))
    if not digits:
        return ""
    return f"〒{digits[:7]}"


def build_addr(c):
    return "".join(
        filter(
            None,
            [
                pref_label(c.get("c.state")),
                t(c.get("c.city")),
                t(c.get("c.town")),
                t(c.get("c.address")),
                t(c.get("c.tatemono")),
            ],
        )
    )


def parse_label_num(name):
    m = re.match(r"^label(\d+)$", str(name), re.IGNORECASE)
    if not m:
        return 999999
    return int(m.group(1))


def resolve_template_path():
    here = Path(__file__).resolve().parent
    candidates = [
        Path("/var/task/assets/template.xlsx"),
        here / "assets" / "template.xlsx",
        here.parent / "assets" / "template.xlsx",
        here.parent.parent / "assets" / "template.xlsx",
        here / "template.xlsx",
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise RuntimeError("template.xlsx がバンドル内に見つかりません（Vercel の includeFiles を確認してください）")


def build_xlsx_bytes(customers):
    if not isinstance(customers, list) or not customers:
        raise RuntimeError("customers is empty")
    template_path = resolve_template_path()
    wb = load_workbook(template_path)
    label_names = sorted(
        [n for n in wb.sheetnames if re.match(r"^label\d+$", str(n), re.IGNORECASE)],
        key=parse_label_num,
    )
    if not label_names:
        label_names = list(wb.sheetnames)

    sheets_needed = int(math.ceil(len(customers) / 12.0))
    if len(label_names) < sheets_needed:
        raise RuntimeError(
            f"テンプレートのlabelシートが不足: required={sheets_needed}, actual={len(label_names)}"
        )

    for i in range(sheets_needed):
        ws = wb[label_names[i]]
        start = i * 12
        chunk = customers[start : start + 12]
        for idx, slot in enumerate(SLOTS):
            c = chunk[idx] if idx < len(chunk) else None
            if not c:
                ws[slot["zip"]].value = ""
                ws[slot["addr"]].value = ""
                ws[slot["name"]].value = ""
                ws[slot["sama"]].value = ""
                continue
            ws[slot["zip"]].value = format_zip(c.get("c.zip"))
            ws[slot["addr"]].value = build_addr(c)
            ws[slot["name"]].value = t(c.get("c.name"))
            ws[slot["sama"]].value = "様"

    for i in range(len(label_names) - 1, sheets_needed - 1, -1):
        wb.remove(wb[label_names[i]])

    for i in range(sheets_needed):
        wb.worksheets[i].title = f"label{i + 1}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class handler(BaseHTTPRequestHandler):
    def log_message(self, _format, *_args):
        return

    def do_POST(self):
        try:
            expected = os.environ.get("REM_LABEL_PRINT_INTERNAL_SECRET", "")
            got = self.headers.get("x-rem-label-secret", "")
            if not expected or got != expected:
                self.send_response(401)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.end_headers()
                self.wfile.write(b'{"ok":false,"message":"unauthorized"}')
                return

            length = int(self.headers.get("Content-Length", "0") or "0")
            raw = self.rfile.read(length) if length > 0 else b"{}"
            body = json.loads(raw.decode("utf-8"))
            customers = body.get("customers") or []
            data = build_xlsx_bytes(customers)
            self.send_response(200)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            msg = str(e)
            self.send_response(500)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.end_headers()
            self.wfile.write(
                json.dumps({"ok": False, "message": msg}, ensure_ascii=False).encode("utf-8")
            )
