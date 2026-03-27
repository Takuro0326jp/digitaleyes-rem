import json
import math
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


PREFS = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県", "茨城県", "栃木県", "群馬県",
    "埼玉県", "千葉県", "東京都", "神奈川県", "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県", "福岡県",
    "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]

SLOTS = [
    {"zip": "B2", "addr": "B3", "name": "C8", "sama": "D8"},
    {"zip": "G2", "addr": "G3", "name": "H8", "sama": "I8"},
    {"zip": "B12", "addr": "B13", "name": "C18", "sama": "D18"},
    {"zip": "G12", "addr": "G13", "name": "H18", "sama": "I18"},
    {"zip": "B22", "addr": "B23", "name": "C28", "sama": "D28"},
    {"zip": "G22", "addr": "G23", "name": "H28", "sama": "I28"},
    {"zip": "B32", "addr": "B33", "name": "C38", "sama": "D38"},
    {"zip": "G32", "addr": "G33", "name": "H38", "sama": "I38"},
    {"zip": "B42", "addr": "B43", "name": "C48", "sama": "D48"},
    {"zip": "G42", "addr": "G43", "name": "H48", "sama": "I48"},
    {"zip": "B52", "addr": "B53", "name": "C58", "sama": "D58"},
    {"zip": "G52", "addr": "G53", "name": "H58", "sama": "I58"},
]


def t(v):
    if v is None:
        return ""
    return str(v).strip()


def pref_label(raw):
    s = t(raw)
    if not s:
        return ""
    if s in PREFS:
        return s
    try:
        n = float(s)
        i = int(n)
        if 1 <= i <= len(PREFS):
            return PREFS[i - 1]
    except Exception:
        pass
    return s


def format_zip(raw):
    digits = re.sub(r"\D", "", t(raw))
    if not digits:
        return ""
    return f"〒{digits[:7]}"


def build_addr(c):
    return "".join(
        filter(
            None,
            [
                pref_label(c.get("c.state")),
                t(c.get("c.city")),
                t(c.get("c.town")),
                t(c.get("c.address")),
                t(c.get("c.tatemono")),
            ],
        )
    )


def parse_label_num(name):
    m = re.match(r"^label(\d+)$", str(name), re.IGNORECASE)
    if not m:
        return 999999
    return int(m.group(1))


def main():
    if len(sys.argv) < 4:
        raise RuntimeError("usage: label-print-openpyxl.py <template.xlsx> <input.json> <output.xlsx>")
    template_path = Path(sys.argv[1])
    input_json_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    payload = json.loads(input_json_path.read_text(encoding="utf-8"))
    customers = payload.get("customers") or []
    if not isinstance(customers, list) or not customers:
        raise RuntimeError("customers is empty")

    wb = load_workbook(template_path)
    label_names = sorted([n for n in wb.sheetnames if re.match(r"^label\d+$", str(n), re.IGNORECASE)], key=parse_label_num)
    if not label_names:
        label_names = list(wb.sheetnames)

    sheets_needed = int(math.ceil(len(customers) / 12.0))
    if len(label_names) < sheets_needed:
        raise RuntimeError(f"テンプレートのlabelシートが不足しています: required={sheets_needed}, actual={len(label_names)}")

    # 先頭Nシートに埋める（既存書式を保持）
    for i in range(sheets_needed):
        ws = wb[label_names[i]]
        start = i * 12
        chunk = customers[start:start + 12]
        for idx, slot in enumerate(SLOTS):
            c = chunk[idx] if idx < len(chunk) else None
            if not c:
                ws[slot["zip"]].value = ""
                ws[slot["addr"]].value = ""
                ws[slot["name"]].value = ""
                ws[slot["sama"]].value = ""
                continue
            ws[slot["zip"]].value = format_zip(c.get("c.zip"))
            ws[slot["addr"]].value = build_addr(c)
            ws[slot["name"]].value = t(c.get("c.name"))
            ws[slot["sama"]].value = "様"

    # 不要なlabelシートを削除
    for i in range(len(label_names) - 1, sheets_needed - 1, -1):
        wb.remove(wb[label_names[i]])

    # シート名を label1..labelN に統一
    for i in range(sheets_needed):
        wb.worksheets[i].title = f"label{i + 1}"

    wb.save(output_path)


if __name__ == "__main__":
    main()

